import pandas as pd
import numpy as np
import operator
import warnings
import re
import os
import openpyxl

# Отключаем warnings
warnings.simplefilter("ignore")


class Decomposition(object):
    """
    Класс с методами необходимыми для декомпозиции.

    Пользовательские методы:
    ------------------------
    - load_WIOD2013_merged_data(**path_and_sheetnames) - чтение данных из таблиц WIOD release 2013. Таблицы
    отечественного выпуска и импорта находятся на одной странице, данные за разные годы лежат в одном файле.
    - load_Rosstat_separated_data(**path_and_sheetnames) - чтение данных из таблиц Росстата. Таблицы отечественного
    выпуска и импорта находятся на разных страницах, данные за разные годы лежат в разных файлах.
    - prepare_data(column_order) - подготовка данных перед декомпозицией

    Методы декомпозиции:
    --------------------
    - decomposition_Baranov_2016() - декомпозиция из статьи в "Вопросах статистики" 2016 год по 2м факторам.
    - decomposition_Baranov_2018() - декомпозиция по статье "Методологические проблемы использования метода структурной
    декомпозиции в модели "затраты – выпуск" на современном этапе" (2018) - по 6 факторам - промежуточный спрос,
    внешний спрос, спрос домашних хозяйств, спрос государства, валовое накопление основного капитала,
    изменение материальных оборотных средств
    - decomposition_Magacho_2018() - декомпозиция из статьи Magacho, G.R., et al., "Impacts of trade liberalization
    on countries’ sectoral structure of production and trade: A structural decomposition analysis." Structural Change
    and Economics Dynamics (2018)

    Методы используемые для внутренних операций:
    --------------------------------------------
    - get_by_name(name) - возвращает столбец по номеру столбца с конца из матриц обоих годов для отечественной
    продукции и импорта
    - get_table(df) - возвращает матрицы затрат на покупку продуктов одних отраслей для производства продуктов других
    отраслей
    - check_sums_equality(self, Z, F, Total) - проверяет правильность полученной таблицы X или M с точностью до 10^-8
    - save_to_excel(self, file_name, rounding="%.3f", **tables) - сохраняет полученные таблицы в выбранный excel-файл
    - pack_name(str) - приклеивает к строке годы, к которым относится содержание строки, и если было указано, то год,
    для которого были пересчитаны цены в таблице

    """

    def __init__(self):
        self.prices_in = ""  # в каких ценах пересчитаны таблицы
        self.years = []  # годы за которые приведены таблицы
        self.df_d = []  # таблицы отечественного выпуска
        self.df_m = []  # таблицы импорта

    def load_WIOD2013_merged_data(self, **path_and_sheetnames):
        """
        Чтение данных из таблиц WIOD release 2013. Таблицы отечественного выпуска и импорта находятся на одной
        странице, данные за разные годы лежат в одном файле.

        Parameters
        ----------
        path_and_sheetnames: dictionary
            путь к excel-файлу и страницы в excel-файле

        """

        # Расположение таблицы и столбцов\строк с названиями в ней
        vertical_table_position = slice(5, 75)  # положение и размеры таблицы по вертикали
        horizontal_table_position = slice(4, 46)  # положение и размеры таблицы по горизонтали
        industries_part_position = slice(4, 39)  # положение и размеры части таблицы с промежуточным потреблением по
        # горизонтали
        codes_position = 1  # номер строки в таблице с кодами отраслей
        columns_names_position = 2  # номер строки в таблице с названиями колонок
        rows_names_position = 1  # номер столбца в таблице с названиями строк


        df_all = []
        for path, sheetnames in path_and_sheetnames.items():
            file = pd.ExcelFile(path)
            for sheetname in sheetnames:
                df1 = pd.read_excel(file, sheet_name=sheetname)

                # Получаем имена столбцов и строк
                rows = df1.iloc[vertical_table_position, rows_names_position]
                rows.name = ""
                columns = df1.iloc[columns_names_position, horizontal_table_position]
                columns.name = df1.columns[0]

                # Получаем из таблицы коды отраслей (одинаковые по вертикали и горизонтали)
                self.codes = df1.iloc[codes_position, industries_part_position]

                # Сохраняем обрезанную версию таблицы
                df = df1.iloc[vertical_table_position, horizontal_table_position]
                df.columns = columns
                df.index = rows
                df.name = columns.name

                df_all.append(df)
                self.years.append(re.search("\d+", columns.name).group(0))

            print("Обрабатываем данные из таблицы \"" + columns.name + "\"")
            self.prices_in = "(в ценах " + re.search("\d+", df1.iloc[0, 0]).group(0) + "года)"

            # Делим таблицы на отечественный выпуск и импорт
            self.df_d = [df_all[0].iloc[:35], df_all[1].iloc[:35]]
            self.df_m = [df_all[0].iloc[35:], df_all[1].iloc[35:]]

    def load_Rosstat_separated_data(self, **path_and_sheetnames):
        """
        Чтение данных из таблиц Росстата. Таблицы отечественного выпуска и импорта находятся на разных страницах,
        данные за разные годы лежат в разных файлах.

        Parameters
        ----------
        path_and_sheetnames: dictionary
            путь к excel-файлу и страницы в excel-файле

        """

        # Расположение таблицы и столбцов\строк с названиями в ней
        vertical_table_position = slice(3, 62)  # положение и размеры таблицы по вертикали
        horizontal_table_position = slice(3, 69)  # положение и размеры таблицы по горизонтали
        industries_part_position = slice(3, 62)  # положение и размеры части таблицы с промежуточным потреблением по
        # горизонтали
        codes_position = 1  # номер строки в таблице с кодами отраслей
        columns_names_position = 0  # номер строки в таблице с названиями колонок
        rows_names_position = 2  # номер столбца в таблице с названиями строк


        for path, sheetnames in path_and_sheetnames.items():
            file = pd.ExcelFile(path)
            for i, sheetname in enumerate(sheetnames):
                df1 = pd.read_excel(file, sheet_name=sheetname)

                # Получаем имена столбцов и строк
                rows = df1.iloc[vertical_table_position, rows_names_position]
                rows.name = ""
                columns = df1.iloc[columns_names_position, horizontal_table_position]
                columns.name = df1.columns[0]

                # Получаем из таблицы коды отраслей (одинаковые по вертикали и горизонтали)
                self.codes = df1.iloc[codes_position, industries_part_position]

                # Сохраняем обрезанную версию таблицы
                df = df1.iloc[vertical_table_position, horizontal_table_position]
                df.columns = columns
                df.index = rows
                df.name = columns.name

                # Делим таблицы на отечественный выпуск и импорт
                if not i:
                    self.df_d.append(df)
                else:
                    self.df_m.append(df)

            print("Обрабатываем данные из таблицы \"" + columns.name + "\"")
            self.years.append(re.search("\d+", columns.name).group(0))

    def get_by_pos(self, pos):
        """
        Возвращает столбец по номеру столбца с конца из матриц обоих годов для отечественной продукции и импорта
        """
        get_by_pos0 = lambda df: df.iloc[:, pos]
        res_m = np.array(list(map(get_by_pos0, self.df_m)))
        res_d = np.array(list(map(get_by_pos0, self.df_d)))
        return res_m, res_d

    def get_table(self, df):
        """
        Возвращает матрицы затрат на покупку продуктов одних отраслей для производства продуктов других отраслей
        """
        return df.iloc[:, :-7]

    def pack_name(self, str):
        """
        Приклеивает к строке годы, к которым относится содержание строки, и если было указано, то год, для которого были
        пересчитаны цены в таблице

        str: string
            строка-название таблицы\страницы файла
        """
        return str + " за " + self.my_years + "гг " + self.prices_in

    def check_sums_equality(self, Z, F, Total):
        """
        Проверяет правильность полученной таблицы X или M с точностью до 10^-8
        """
        for year in [0, 1]:
            z = Z[year]
            f = F[year]
            total = Total[year]
            for i in range(len(total)):
                if (abs(sum(z.iloc[i]) + f[i] - total[i]) >= 1e-8):
                    print("Ошибка в таблице \"", z.columns.name, "\" в строке \"", z.columns[i],
                          "\" должно быть:", sum(z.iloc[i]) + f[i], ", написано:", total[i])
                    Total[year][i] = sum(z.iloc[i]) + f[i]
        return Total

    def prepare_data(self, column_order):
        """
        Подготовка данных перед декомпозицией

        column_order: string
            Порядок столбцов в таблице "eng" - households, NPISH, government or "rus" - households, government, NPISH
        """
        self.columns = self.df_d[0].columns.values
        self.my_years = self.years[0] + "-" + self.years[1]
        os.makedirs("./results/" + self.my_years, exist_ok=True)

        if column_order == 'eng':
            self.C_mnpish, self.C_dnpish = self.get_by_pos(-6)
            self.C_mg, self.C_dg = self.get_by_pos(-5)
        else:
            self.C_mg, self.C_dg = self.get_by_pos(-6)
            self.C_mnpish, self.C_dnpish = self.get_by_pos(-5)

        self.C_mh, self.C_dh = self.get_by_pos(-7)
        self.I_mgfch, self.I_dgfch = self.get_by_pos(-4)
        self.I_ms, self.I_ds = self.get_by_pos(-3)
        self.E_r, self.E = self.get_by_pos(-2)
        self.M, self.X = self.get_by_pos(-1)

        Z_m = list(map(self.get_table, self.df_m))
        Z_d = list(map(self.get_table, self.df_d))

        self.X[self.X == 0] = 1e-20
        self.M[self.M == 0] = 1e-20

        # Вычисляем суммы столбцов таблицы
        self.I_m = self.I_mgfch + self.I_ms
        self.I_d = self.I_dgfch + self.I_ds
        self.C_mh = self.C_mh + self.C_mnpish
        self.C_dh = self.C_dh + self.C_dnpish
        self.C_m = self.C_mg + self.C_mh
        self.C_d = self.C_dg + self.C_dh
        self.F_m = self.I_m + self.C_m + self.E_r
        self.F_d = self.I_d + self.C_d + self.E

        # Проверяем правильность столбцов суммарного выпуска
        self.M = self.check_sums_equality(Z_m, self.F_m, self.M)
        self.X = self.check_sums_equality(Z_d, self.F_d, self.X)

        # Вычисляем матрицы технических коэффициентов
        self.A_d = [(Z_d[0] / self.X[0]).astype('float'), (Z_d[1] / self.X[1]).astype('float')]
        self.A_m = [Z_m[0] / self.X[0], Z_m[1] / self.X[1]]

        I = np.eye(np.size(self.X[0]), dtype='float')
        self.L_d = [np.linalg.inv(((I - A).astype('float'))) for A in self.A_d]

    def decomposition_Baranov_2016(self):
        """
        Метод декомпозиции из статьи в "Вопросах статистики" 2016 год. Декомпозиция по 2м факторам.

        """
        columns = ['dX полученный с помощью метода декомпозиции, млн', 'Разность X1 - X0, млн',
                   'dM полученный с помощью метода декомпозиции, млн', 'Разность M1 - M0, млн']
        columns_perc = ['Выпуск ' + self.years[0] + ' года, млн', 'Выпуск ' + self.years[1] + ' года, млн',
                        'Выпуск ' + self.years[1] + ' года к ' + self.years[0] + ' году, %',
                        'Импорт ' + self.years[0] + ' года, млн', 'Импорт ' + self.years[1] + ' года, млн',
                        'Импорт ' + self.years[1] + ' года к ' + self.years[0] + ' году, %']

        WL0 = self.A_m[0].dot(self.L_d[0])
        WL1 = self.A_m[1].dot(self.L_d[1])

        dX = ((self.L_d[1] + self.L_d[0]).dot(self.F_d[1] - self.F_d[0]) + (self.L_d[1] - self.L_d[0]).dot(
            self.F_d[1] + self.F_d[0])) / 2

        dM = ((WL1 + WL0).dot(self.F_d[1] - self.F_d[0]) + (WL1 - WL0).dot(self.F_d[1] + self.F_d[0])) / 2 + \
             self.F_m[1] - self.F_m[0]

        dX_perc = (self.X[1] / self.X[0]) * 100
        dM_perc = (self.M[1] / self.M[0]) * 100

        # Вывод таблиц
        results = pd.DataFrame(np.column_stack([dX, self.X[1] - self.X[0], dM, self.M[1] - self.M[0]]),
                               columns=columns, index=self.df_d[0].index)
        results.loc["Total"] = [sum(dX), sum(self.X[1]) - sum(self.X[0]), sum(dM), sum(self.M[1]) - sum(self.M[0])]
        results.columns.name = 'Изменения в выпуске и импорте'

        self.results_percents = pd.DataFrame(
            np.column_stack([self.X[0], self.X[1], dX_perc, self.M[0], self.M[1], dM_perc]),
            columns=columns_perc,
            index=self.df_d[0].index)
        self.results_percents.loc["Total"] = [sum(self.X[0]), sum(self.X[1]), (sum(self.X[1]) / sum(self.X[0])) * 100,
                                              sum(self.M[0]), sum(self.M[1]), (sum(self.M[1]) / sum(self.M[0])) * 100]
        self.results_percents.columns.name = 'Изменения в выпуске и импорте'

        # Проверяем, что изменения в валовом выпуске, полученные как сумма факторов (dX и dM) сходятся с разностями
        # X[1] - X[0] и M[1] - M[0], полученными из таблиц (с точностью до 10^-5)
        assert (sum(self.X[1]) - sum(self.X[0]) - sum(
            dX) < 1e-5), "Oops! Полученные суммарные изменения в валовом выпуске dX не равны X1 - X0!"
        assert (sum(self.M[1]) - sum(self.M[0]) - sum(
            dM) < 1e-5), "Oops! Полученные суммарные изменения в валовом выпуске dM не равны M1 - M0!"

        result_tables = {
            "Упрощенная декомпозиция изменений в выпуске и импорте за " + self.my_years + "гг " + self.prices_in: results}
        self.save_to_excel('results_simple(Baranov_2016).xlsx', **result_tables)

        print("\nРезультат работы метода декомпозиции Baranov_2016 сохранен в папку results!")

    def decomposition_Baranov_2018(self):
        """
        Метод декомпозиции из статьи "Методологические проблемы использования метода структурной декомпозиции
        в модели "затраты – выпуск" на современном этапе" (2018).

        Декомпозиция по 6 факторам - промежуточный спрос, внешний спрос, спрос домашних хозяйств, спрос
        государства, валовое накопление основного капитала, изменение материальных оборотных средств

        """
        # Колонки для результирующих таблиц
        columns_X = ['Конечный спрос через формулу для декомпозиции 16 года',
                  'Конечный спрос как сумма факторов', 'dX полученный с помощью метода декомпозиции',
                  'Разность X1 - X0']
        columns_M = ['Конечный спрос через формулу для декомпозиции 16 года',
                  'Конечный спрос как сумма факторов', 'dM полученный с помощью метода декомпозиции',
                  'Разность M1 - M0']
        res_columns = ['Промежуточный спрос', 'Внешний спрос', 'Спрос домашних хозяйств', 'Спрос государства',
                    'Валовое накопление основного капитала', 'Изменение запаса материальных оборотных средств']
        res_index = ['Выпуск отечественной продукции', 'Импорт', 'Всего']
        res_columns_perc = ['Промежуточный спрос, %', 'Внешний спрос, %', 'Спрос домашних хозяйств, %',
                         'Спрос государства, %', 'Валовое накопление основного капитала, %',
                         'Изменение запаса материальных оборотных средств, %']

        sumL = self.L_d[0] + self.L_d[1]
        sumW = self.A_m[0].dot(self.L_d[0]) + self.A_m[1].dot(self.L_d[1])

        # Получаем слагаемые декомпозиции изменения выпуска отечественной продукции
        dX = np.array([np.zeros(len(self.M[0]))] * 8)
        dX[0] = ((self.L_d[1] - self.L_d[0]).dot(self.F_d[1] + self.F_d[0])) / 2
        dX[1] = sumL.dot(self.E[1] - self.E[0]) / 2
        dX[2] = sumL.dot(self.C_dh[1] - self.C_dh[0]) / 2
        dX[3] = sumL.dot(self.C_dg[1] - self.C_dg[0]) / 2
        dX[4] = sumL.dot(self.I_dgfch[1] - self.I_dgfch[0]) / 2
        dX[5] = sumL.dot(self.I_ds[1] - self.I_ds[0]) / 2
        dX_all = dX[0] + dX[1] + dX[2] + dX[3] + dX[4] + dX[5]
        Xtot = sum(dX_all)

        dX[6] = dX[1] + dX[2] + dX[3] + dX[4] + dX[5]
        dX[7] = ((self.L_d[1] + self.L_d[0]).dot(self.F_d[1] - self.F_d[0])) / 2

        # Получаем слагаемые декомпозиции изменения импорта
        dM = np.array([np.zeros(len(self.M[0]))] * 8)
        dM[0] = ((self.A_m[1] - self.A_m[0]).dot(self.L_d[1].dot(self.F_d[1]) + self.L_d[0].dot(self.F_d[0])) +
                 self.A_m[0].dot(self.L_d[1] - self.L_d[0]).dot(self.F_d[1]) + self.A_m[1].dot(
            self.L_d[1] - self.L_d[0]).dot(self.F_d[0])) / 2

        dM[1] = self.E_r[1] - self.E_r[0] + sumW.dot(self.E[1] - self.E[0]) / 2
        dM[2] = self.C_mh[1] - self.C_mh[0] + sumW.dot(self.C_dh[1] - self.C_dh[0]) / 2
        dM[3] = self.C_mg[1] - self.C_mg[0] + sumW.dot(self.C_dg[1] - self.C_dg[0]) / 2
        dM[4] = self.I_mgfch[1] - self.I_mgfch[0] + sumW.dot(self.I_dgfch[1] - self.I_dgfch[0]) / 2
        dM[5] = self.I_ms[1] - self.I_ms[0] + sumW.dot(self.I_ds[1] - self.I_ds[0]) / 2
        dM_all = dM[0] + dM[1] + dM[2] + dM[3] + dM[4] + dM[5]
        Mtot = sum(dM_all)

        dM[6] = dM[1] + dM[2] + dM[3] + dM[4] + dM[5]
        dM[7] = ((self.A_m[1].dot(self.L_d[1]) + self.A_m[0].dot(self.L_d[0])).dot(self.F_d[1] - self.F_d[0])) / 2 + (
            self.F_m[1] - self.F_m[0])

        # Проверяем, что изменения в валовом выпуске, полученные как сумма факторов (dX_all и dM_all) сходятся с
        # разностями X[1] - X[0] и M[1] - M[0], полученными из таблиц (с точностью до 10^-5)
        assert (sum(self.X[1]) - sum(self.X[0]) - sum(
            dX_all) < 1e-5), "Oops! Полученные суммарные изменения в валовом выпуске dX_all не равны X1 - X0!"
        assert (sum(self.M[1]) - sum(self.M[0]) - sum(
            dM_all) < 1e-5), "Oops! Полученные суммарные изменения в валовом выпуске dM_all не равны M1 - M0!"

        # Вывод таблиц
        result_d = pd.DataFrame(np.column_stack(dX[:6]), columns=res_columns, index=self.df_d[0].index)
        result_d.columns.name = 'Выпуск отечественной продукции'
        result_m = pd.DataFrame(np.column_stack(dM[:6]), columns=res_columns, index=self.df_d[0].index)
        result_m.columns.name = 'Импорт'

        res_check_X = pd.DataFrame(np.column_stack([dX[7], dX[6], dX_all, self.X[1] - self.X[0]]),
                                   columns=columns_X, index=self.df_d[0].index)
        res_check_X.columns.name = 'Выпуск отечественной продукци'
        res_check_M = pd.DataFrame(np.column_stack([dM[7], dM[6], dM_all, self.M[1] - self.M[0]]),
                                   columns=columns_M, index=self.df_d[0].index)
        res_check_M.columns.name = 'Импорт'

        # Вывод таблиц в процентах - таблицы аналогичные полученным в "Вопросах статистики"
        res_perc = pd.DataFrame(
            np.column_stack([(dX[0] + dM[0]) / abs((dX_all + dM_all)) * 100, (dX[6] + dM[6]) / abs((dX_all + dM_all)) *
                             100]),
            columns=["Промежуточный спрос, %", "Конечный спрос, %"], index=self.df_d[0].index)
        res_perc.columns.name = "Изменение спроса суммарно"

        res_perc1 = pd.DataFrame(np.column_stack([dX[0] / abs((dX[0] + dM[0])) * 100, dM[0] / abs((dX[0] + dM[0])) *
                                                  100]),
                                 columns=["Отечественная продукция, %", "Импорт, %"], index=self.df_d[0].index)
        res_perc1.columns.name = "Изменение промежуточного спроса"

        res_perc2 = pd.DataFrame(
            np.column_stack([dX[6] / abs((dX[6] + dM[6])) * 100, dM[6] / abs((dX[6] + dM[6])) * 100]),
            columns=["Отечественная продукция, %", "Импорт, %"], index=self.df_d[0].index)
        res_perc2.columns.name = "Изменение конечного спроса"

        results_perc_d = pd.DataFrame(np.column_stack(dX[:6] / abs(dX_all) * 100), index=self.df_d[0].index,
                                      columns=res_columns)
        results_perc_d.columns.name = 'Изменение отечественого выпуска'

        results_perc_m = pd.DataFrame(np.column_stack(dM[:6] / abs(dM_all) * 100), index=self.df_d[0].index,
                                      columns=res_columns)
        results_perc_m.columns.name = 'Изменение импорта'

        # res_perc.loc['Private Households with Employed Persons'] = [0, 100]
        # results_perc_d.loc['Private Households with Employed Persons'] = [""] * 6
        # results_perc_m.loc['Private Households with Employed Persons'] = [""]*6

        # Суммы
        sumss_d = list(map(sum, dX[:6]))
        sumss_m = list(map(sum, dM[:6]))
        sumss = list(map(operator.add, sumss_d, sumss_m))

        # Добавляем строки с суммой в конец таблиц
        res_check_X.loc['Total'] = [sum(dX[7]), sum(dX[6]), Xtot, sum(self.X[1]) - sum(self.X[0])]
        res_check_M.loc['Total'] = [sum(dM[7]), sum(dM[6]), Mtot, sum(self.M[1]) - sum(self.M[0])]
        result_d.loc['Total'] = sumss_d
        result_m.loc['Total'] = sumss_m

        res_perc.loc['Total'] = [np.round(sum(dX[0] + dM[0]) / abs(sum(dX_all + dM_all)) * 100, 1),
                                 np.round(sum(dX[6] + dM[6]) / abs(sum(dX_all + dM_all)) * 100, 1)]
        res_perc1.loc['Total'] = [np.round(sum(dX[0]) / abs(sum(dX[0] + dM[0])) * 100, 1),
                                  np.round(sum(dM[0]) / abs(sum(dX[0] + dM[0])) * 100, 1)]
        res_perc2.loc['Total'] = [np.round(sum(dX[6]) / abs(sum(dX[6] + dM[6])) * 100, 1),
                                  np.round(sum(dM[6]) / abs(sum(dX[6] + dM[6])) * 100, 1)]
        results_perc_d.loc['Total'] = np.round(sumss_d / abs(sum(dX_all)) * 100, 1)
        results_perc_m.loc['Total'] = np.round(sumss_m / abs(sum(dM_all)) * 100, 1)

        results = pd.DataFrame([sumss_d, sumss_m, sumss], index=res_index, columns=res_columns)
        results.columns.name = 'Изменение выпуска за ' + self.years[0] + "-" + self.years[1]

        results_perc = pd.DataFrame(
            [sumss_d / abs(sum(sumss_d)) * 100, sumss_m / abs(sum(sumss_m)) * 100, sumss / abs(sum(sumss)) * 100],
            index=res_index, columns=res_columns_perc)
        results_perc.columns.name = 'Изменение выпуска за ' + self.years[0] + "-" + self.years[1]



        # Присваиваем имена таблицам и сохраняем в excel
        percented_result_tables = {self.pack_name("Декомпозиция изменений выпуска по всем факторам"): results_perc,
                                   self.pack_name("Декомпозиция изменений отечественного выпуска по всем факторам и "
                                   "по всем отраслям"): results_perc_d,
                                   "Декомпозиция изменений импорта по всем факторам и по всем отраслям за "
                                   + self.my_years + "гг " + self.prices_in: results_perc_m}

        percented_tables = {"Изменения в выпуске и импорте за " + self.my_years + "гг "
                            + self.prices_in: self.results_percents,
                            "Изменение спроса суммарно для отечественной продукции и "
                            "импорта в процентах за " + self.my_years + "гг " + self.prices_in: res_perc,
                            "Изменение промежуточного спроса в процентах за " + self.my_years +
                            "гг " + self.prices_in: res_perc1,
                            "Изменение конечного спроса в процентах за " + self.my_years +
                            "гг " + self.prices_in: res_perc2}

        result_tables = {'Декомпозиция изменений выпуска по всем факторам за ' + self.my_years +
                         'гг ' + self.prices_in: results,
                         'Выпуск отечественной продукци за ' + self.my_years +
                         'гг ' + self.prices_in: result_d,
                         'Импорт за ' + self.my_years +
                         'гг ' + self.prices_in: result_m}
        checking_tables = {'Выпуск отечественной продукци за ' + self.my_years + 'гг ' + self.prices_in: res_check_X,
                           'Импорт за ' + self.my_years + 'гг ' + self.prices_in: res_check_M}

        self.save_to_excel('results_in_percents(Baranov_2018).xlsx',
                           **percented_result_tables)
        self.save_to_excel('changes_in_percents.xlsx', **percented_tables)
        self.save_to_excel('results(Baranov_2018).xlsx', **result_tables)
        self.save_to_excel('для_проверки_(Baranov_2018).xlsx', **checking_tables)

        print("Результат работы метода декомпозиции Baranov_2018 сохранен в папку results!\n")

    def decomposition_Magacho_2018(self):
        """
        Метод декомпозиции из статьи Magacho, G.R., et al., "Impacts of trade liberalization on countries’ sectoral
        structure of production and trade: A structural decomposition analysis." Structural Change and Economics
        Dynamics (2018)

        Декомпозиция по 3 факторам - технологические изменения, замещение национальных продуктов
        импортированными(?), конечный спрос

        """
        # Колонки для результирующих таблиц
        res_columns = ['Технологические изменения', 'Замещение отечественных продуктов импортированными',
                       'Конечный спрос(включая экспорт)', 'Экспорт', 'dX полученный с помощью метода декомпозиции',
                       'Разность X1 - X0']

        dA = self.A_m[1] + self.A_d[1] - self.A_m[0] - self.A_d[0]
        sumF = self.F_d[1] + self.F_d[0]

        # Получаем слагаемые декомпозиции изменения выпуска отечественной продукции
        technological_change = (self.L_d[1].dot(dA)).dot(self.L_d[0]).dot(sumF) / 2
        substitution_national_inputs = (self.L_d[1].dot(self.A_m[0] - self.A_m[1])).dot(self.L_d[0]).dot(sumF) / 2
        final_demands = (self.L_d[1] + self.L_d[0]).dot(self.F_d[1] - self.F_d[0]) / 2
        export = (self.L_d[1] + self.L_d[0]).dot(self.E[1] - self.E[0]) / 2
        dX = technological_change + substitution_national_inputs + final_demands

        dX[dX == 0] = 1e-20

        #dX_perc = (self.X[1] / self.X[0]) * 100

        # Проверяем, что изменения в валовом выпуске, полученные как сумма факторов (dX) сходятся с разностью
        # X[1] - X[0], полученной из таблиц (с точностью до 10^-5)
        assert (sum(self.X[1]) - sum(self.X[0]) - sum(dX) < 1e-5), \
            "Oops! Полученные суммарные изменения в валовом выпуске dX не равны X1 - X0!"

        # Заполняем таблицы
        results = pd.DataFrame(np.column_stack([technological_change, substitution_national_inputs,
                                                       final_demands, export,
                                                dX, self.X[1] - self.X[0]]), index=self.df_d[0].index,
                               columns= res_columns)
        results.columns.name = 'Выпуск отечественной продукции'


        results_percented = pd.DataFrame(np.column_stack([technological_change / abs(dX) * 100,
                                                        substitution_national_inputs / abs(dX) * 100,
                                                        final_demands / abs(dX) * 100, export / abs(dX) * 100]),
                                         index=self.df_d[0].index,
                               columns=res_columns[:-2])
        results_percented.columns.name = 'Выпуск отечественной продукции'


        results.loc["Total"] = [sum(technological_change), sum(substitution_national_inputs), sum(final_demands),
                                sum(export), sum(dX), sum(self.X[1]) - sum(self.X[0])]
        results_percented.loc["Total"] = [sum(technological_change) / abs(sum(dX)) * 100,
                                          sum(substitution_national_inputs) / abs(sum(dX)) * 100,
                                         sum(final_demands) / abs(sum(dX)) * 100,
                                          sum(export) / abs(sum(dX)) * 100]



        # Присваиваем имена таблицам и сохраняем в excel
        result_tables = {self.pack_name('Выпуск отечественной продукции'): results}
        result_percented_tables = {self.pack_name('Выпуск отечественной продукции'): results_percented}
        self.save_to_excel('results(Magacho_2018).xlsx', **result_tables)
        self.save_to_excel('results_in_percents(Magacho_2018).xlsx', **result_percented_tables)

        print("Результат работы метода декомпозиции Magacho_2018 сохранен в папку results!\n")

    def save_to_excel(self, file_name, rounding="%.3f", **tables):
        """
        Cохраняет полученные таблицы в выбранный excel-файл

        Parameters
        ----------
        file_name: string
            имя файла
        rounding: string
            Формат округления чисел в сохраняемых таблицах.
            по умолчанию - "%.3f"(до 3 знаков после запятой),
            для результатов в процентах - "%.1f"(до 1 знака после запятой)
        tables: dictionary
            Словарь из названий таблиц и самих таблиц

        """
        writer = pd.ExcelWriter("./results/" + self.my_years + "/" + file_name, engine='xlsxwriter')
        workbook = writer.book

        for table_name, df in tables.items():
            sheet_name = df.columns.name

            df.to_excel(writer, sheet_name=sheet_name, float_format=rounding, startrow=1, startcol=2,
                        header=False, index=False)

            worksheet = writer.sheets[sheet_name]
            worksheet.set_zoom(80)
            worksheet.set_column(2, len(df.columns) + 1, 20)

            # Add a header format
            header_format = workbook.add_format({
                'bold': True,
                'text_wrap': True,
                'align': 'center',
                'valign': 'vcenter',
                'fg_color': '#D7E4BC',
                'border': 1})

            # Add total row formating
            if any(df.index.values == 'Total'):
                total_format = workbook.add_format({'bold': True, 'fg_color': '#ffb74d', 'border': 1})
                worksheet.write_row(np.shape(df)[0], 2, np.array(df.loc['Total', :]), total_format)

            # Add codes & index
            if len(df.index) > len(self.codes):
                codes2 = pd.concat([pd.Series(['']), self.codes])
                worksheet.write_column('A1', codes2, header_format)
            worksheet.write_column('B1', np.insert(df.index.values, 0, ''), header_format)
            worksheet.set_column('B:B', 45)

            # Add header
            columns = np.insert(df.columns.values, 0, table_name)
            for col_num, value in enumerate(columns):
                worksheet.write(0, col_num + 1, value, header_format)
            worksheet.set_row(0, 60)

        writer.save()
        workbook.close()




dec = Decomposition()

path_and_sheetnames = {"./data/задание 1 для студентов.xlsx": [0, 1]}
dec.load_WIOD2013_merged_data(**path_and_sheetnames)
dec.prepare_data(column_order='eng')

dec.decomposition_Baranov_2016()
dec.decomposition_Baranov_2018()
dec.decomposition_Magacho_2018()

dec = Decomposition()
path_and_sheetnames = {"./data/all2011.xlsx": ['SD calculated def', 'SM calculated def'],
                       "./data/all2014 (проверочная).xlsx": ['SD calculated def', 'SM calculated def']}
dec.load_Rosstat_separated_data(**path_and_sheetnames)
dec.prepare_data(column_order='rus')

dec.decomposition_Baranov_2016()
dec.decomposition_Baranov_2018()
dec.decomposition_Magacho_2018()


dec = Decomposition()
path_and_sheetnames = {"./data/all2014 (проверочная).xlsx": ['SD calculated def', 'SM calculated def'],
                       "./data/all2015 (проверочная).xlsx": ['SD calculated def', 'SM calculated def']}
dec.load_Rosstat_separated_data(**path_and_sheetnames)
dec.prepare_data(column_order='rus')

dec.decomposition_Baranov_2016()
dec.decomposition_Baranov_2018()
dec.decomposition_Magacho_2018()
